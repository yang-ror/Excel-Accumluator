import os
import sys
import shutil
import openpyxl
import colorama
from datetime import datetime
from colorama import Fore, Style
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QScrollArea, QCheckBox, QPushButton, QLabel


colorama.init()
template_file = './template.xlsx'
cell_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

# target_file = './订单物料需求.xlsx'

# def main():
#     target_file = './订单物料需求.xlsx'
#     recorded_orders = get_recorded_orders(target_file)
#     # print(recorded_orders)
    
#     customers = get_all_directories('./')
#     customers = [s for s in customers if 'do not process' not in s]
#     # print(customers)
    
#     work_orders = []
#     for customer in customers:
#         orders = get_orders(customer, recorded_orders)
#         # print(orders)
#         materials = None
#         for order in orders:
#             materials = get_materials(order["path"])
#             # print(materials)
#             if not materials:
#                 print(f'No "materials" or "total" found in {order["path"]}')
#                 continue
#             work_orders.append({
#                 'order_num': order['num'],
#                 # 'title': get_order_num_from_file_name(os.path.splitext(order)[0]),
#                 'materials': materials
#             })
#     # print(work_orders)
    
#     order_list = preapre_orders_for_writing(work_orders)
#     # print(order_list)
#     for order in order_list:
#         write_order_to_file(order, target_file)


def get_order_list():
    # recorded_orders = get_recorded_orders(target_file)
    
    # customers = get_all_directories('./')
    # customers = [s for s in customers if 'do not process' not in s]

    customers = [ 'to be processed' ]
    
    all_orders = []
    
    for customer in customers:
        all_orders.extend(get_all_orders(customer))
    
    order_list = []
    
    for order in all_orders:
        # recorded = False
        # if not order['num'] in recorded_orders:
        #     recorded = True
        order_list.append({
            # "checked": recorded,
            "checked": True,
            "label": order['num'],
            **order
        })
    
    return order_list


def get_recorded_orders(file):
    orders = set()
    try:
        workbook = openpyxl.load_workbook(file)
    except:
        print(Fore.CYAN + f'Cannot find {file}, please check current folder' + Style.RESET_ALL)
        return []
    # sheets_to_process = ['raw', 'ingredient', 'bag', 'box', 'oxygen']
    sheets_to_process = ['raw']
    for sheet_name in sheets_to_process:
        worksheet = workbook[sheet_name]
        for cell in worksheet[1]:
            if cell.value not in ['SKU', 'QTY', 'total']:
                orders.add(cell.value)
    return orders


def get_all_directories(dir):
    subdirs = []
    for name in os.listdir(dir):
        path = os.path.join(dir, name)
        if os.path.isdir(path):
            subdirs.append(path)
    return subdirs


def get_orders(customer, recorded_orders):
    excel_files = [f for f in os.listdir(customer) if f.endswith('.xlsx') and not f.startswith('template') and not f.startswith('dnp')]
    # for f in excel_files:
    #     print(f)
    #     print(os.path.splitext(f)[0])
    #     print(get_order_num_from_file_name(os.path.splitext(f)[0]))
    # new_orders = [os.path.join(customer, f) for f in excel_files if get_order_num_from_file_name(os.path.splitext(f)[0]) not in recorded_orders]
    orders = []
    for order in excel_files:
        order_num = get_order_num_from_file_name(os.path.splitext(order)[0])
        if order_num in recorded_orders:
            continue
        orders.append({
            'num': order_num,
            'path': os.path.join(customer, order)
        })
    return orders


def get_all_orders(customer):
    excel_files = [f for f in os.listdir(customer) if f.endswith('.xlsx') and not f.startswith('template') and not f.startswith('dnp')]
    orders = []
    for order in excel_files:
        order_num = get_order_num_from_file_name(os.path.splitext(order)[0])
        orders.append({
            'num': order_num,
            'path': os.path.join(customer, order)
        })
    return orders


def get_order_num_from_file_name(string):
    if '-' not in string:
        return string.strip()
    substring = string.split('-')[0].strip()
    return substring


def get_materials_for_order(order):
    materials = get_materials(order["path"])
    
    if not materials:
        print(Fore.RED + f'No "materials" or "total" found in {order["path"]}' + Style.RESET_ALL)
        return
    
    return{
        'order_num': order['num'],
        'materials': materials
    }


def get_materials(order):
    materials = []
    wb = openpyxl.load_workbook(order, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
        if 'materials' in row_values and 'total' in row_values:
            row_num = row[0].row
            m_col = row_values.index('materials') + 1
            t_col = row_values.index('total') + 1
            break
    else:
        return materials
    
    row_num += 1
    while True:
        m_value = ws.cell(row=row_num, column=m_col).value
        t_value = ws.cell(row=row_num, column=t_col).value
        
        if not m_value:
            break
        
        try:
            t_value = round(t_value, 3)
        except:
            pass
            
        material = {'id': m_value, 'qty': t_value}
        materials.append(material)
        row_num += 1
    return materials


def preapre_orders_for_writing(work_orders):
    order_list = []
    order_list.extend(get_materials_in_cat('raw', work_orders))
    order_list.extend(get_materials_in_cat('ingredient', work_orders))
    order_list.extend(get_materials_in_cat('bag', work_orders))
    order_list.extend(get_materials_in_cat('box', work_orders))
    order_list.extend(get_materials_in_cat('oxygen', work_orders))
    return order_list
    

def get_materials_in_cat(cat, work_orders):
    letter = get_cat_letter(cat)
    new_list = []
    for order in work_orders:
        filtered_materials = [material for material in order['materials'] if material['id'].startswith(letter)]
        if len(filtered_materials) == 0:
            continue
        new_list.append({
            'order_num': order['order_num'],
            'cat': cat,
            'materials':  filtered_materials
        })
    return new_list


def get_cat_letter(cat):
    if cat == 'raw' : return 'A'
    if cat == 'ingredient' : return 'B'
    if cat == 'bag' : return 'C'
    if cat == 'box' : return 'E'
    if cat == 'oxygen' : return 'D'


def write_order_to_file(order, target_file):
    # Open the Excel file and select the worksheet for the given category
    wb = openpyxl.load_workbook(target_file)

    # ws = wb[order["cat"]]
    try:
        ws = wb[order["cat"]]
    except:
        print(Fore.RED + f'Cannot open {order["cat"]}, please check the excel file.' + Style.RESET_ALL)
        return
        
    print(f'{order["order_num"]} - {order["cat"]}')
    
    # Loop through the first row until we find the 'total' column or an empty cell
    # found_empty_col = False
    for col in range(4, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'total':
            # if not found_empty_col:
            ws.insert_cols(col)
            col_num = col
            ws.cell(row=1, column=col_num).value = order["order_num"]
            break
        elif ws.cell(row=1, column=col).value == order["order_num"] or not ws.cell(row=1, column=col).value:
            # found_empty_col = True
            col_num = col
            # ws.insert_cols(col)
            ws.cell(row=1, column=col_num).value = order["order_num"]
            break
            

    
    # Loop through the materials and write their quantities to the appropriate cells
    for material in order['materials']:
        # Search for the row with the matching material ID
        found_material_id = False
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == material['id']:
                found_material_id = True
                row_num = row
                break
        if not found_material_id:
            print(Fore.RED + f"cannot find {material['id']}, please check the --template-- excel file." + Style.RESET_ALL)
            continue
        # Write the material quantity to the appropriate cell
        
        print(Fore.CYAN + f"{material['id']} - {material['qty']}" + Style.RESET_ALL)
        
        ws.cell(row=row_num, column=col_num).value = material['qty']
        # ws.cell(row=row_num, column=col_num).border = cell_border
    
    total_col = None
    for cell in ws[1]:
        if cell.value == 'total':
            total_col = cell.column
            break

    if total_col is not None:
        for row in range(3, ws.max_row+1):
            total_formula = f"=SUM(D{row}:{openpyxl.utils.get_column_letter(total_col-1)}{row})"
            ws.cell(row=row, column=total_col).value = total_formula

            diff_formula = f"=B{row}-{openpyxl.utils.get_column_letter(total_col)}{row}"
            ws.cell(row=row, column=3).value = diff_formula
    
    # Save the changes to the Excel file
    wb.save(target_file)


class MyApp(QWidget):
    def __init__(self):
        super().__init__()

        order_list = get_order_list()
        # print(order_list)
        # order_list = []
        # for i in range(20):
        #     order_list.append({"checked": True, "label": f'label - {i+1}'})

        # Set the window properties
        self.setGeometry(100, 100, 300, 400)
        self.setWindowTitle('My App')

        # Create the scroll area
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setFixedSize(300, 300)
        self.scroll_area.setWidgetResizable(False)

        self.checkboxes = []
        for order in order_list:
            checkbox = QCheckBox(order['label'])
            checkbox.setChecked(order['checked'])
            checkbox.stateChanged.connect(lambda state, order=order: self.on_checkbox_change(state, order))
            self.checkboxes.append(checkbox)
        
        # Set up the layout
        layout = QVBoxLayout()
        for checkbox in self.checkboxes:
            layout.addWidget(checkbox)
        scroll_widget = QWidget()
        scroll_widget.setLayout(layout)
        self.scroll_area.setWidget(scroll_widget)

        # Create the auto-complete button
        self.button = QPushButton('Auto-complete', self)
        self.button.setStyleSheet("font-size: 16pt;")
        self.button.setFixedSize(300, 100)
        self.button.clicked.connect(lambda: self.proceed(order_list))

        # Set up the main layout
        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.scroll_area)
        self.main_layout.addWidget(self.button)
        self.setLayout(self.main_layout)


    def on_checkbox_change(self, state, order):
        order['checked'] = (state == 2)  # 2 = Checked


    def proceed(self, order_list):
        orders = []
        
        for order in order_list:
            if not order['checked']: continue
            orders.append(get_materials_for_order(order))
        
        orders_to_write = preapre_orders_for_writing(orders)
        
        now = datetime.now()
        time = now.strftime("%m-%d_%H-%M-%S")
        src_file = template_file
        dst_file = f'./订单物料需求{time}.xlsx'
        shutil.copy(src_file, dst_file)
        
        for order in orders_to_write:
           write_order_to_file(order, dst_file)
            
        label = QLabel("Complete")
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("font-size: 16pt;")
        self.scroll_area.setParent(None)
        self.button.setParent(None)
        self.main_layout.addWidget(label)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyApp()
    window.show()
    sys.exit(app.exec_())
